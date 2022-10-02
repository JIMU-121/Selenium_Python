from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='demoexcel.xlsx')
sh = wb['DemoSheet']

# print(sh['A3'].value)
# print(wb['DemoSheet']['A3'].value)
#
# print(sh.cell(row=3, column=1).value)

row_count = sh.max_row
col_count = sh.max_column

print("Rows : "+str(row_count))
print("Columns : "+str(col_count))

for i in range(1, row_count+1):
    for j in range(1, col_count+1):
        print(sh.cell(row=i, column=j).value, end=' ')
    print('\n')