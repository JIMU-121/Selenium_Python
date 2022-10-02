from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# ws['A1'] = "Souvik Bhattacharya"

# testdata = [['Name', 'City'], ['Manish', 'Melbourne'], ['Rama', 'Bangalore'], ['Sam', 'London']]
# # testdata2 = ["Souvik", "Bhattacharya"]
#
# for data in testdata:
#     ws.append(data)


for i in range(1, 6):
    for j in range(1, 5):
        ws.cell(row=i, column=j).value = i+j


wb.save("demoexcel.xlsx")