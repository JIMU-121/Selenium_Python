import time
import os
import random
import sys
from selenium import webdriver
from selenium.common import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from ExtractImages import FetchingImages
from Products import Fields, Actions, Categories
from Actions import HairDryersAction as Action
from openpyxl import Workbook, load_workbook
from numba import jit




# Files Path
ProImgDir = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\1290 + mini + cp\\images'
SlideImagesPath = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\1290 + mini + cp\\slide'
path = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\1290 + mini + cp'
# SlideImg1 = 'C:/Users/bhatt/Downloads/Telegram Desktop/folder 8/slide/img1.jpg'
# SlideImg2 = 'C:/Users/bhatt/Downloads/Telegram Desktop/folder 8/slide/img2.jpg'
# SlideImg3 = 'C:/Users/bhatt/Downloads/Telegram Desktop/folder 8/slide/OIP.jpeg'

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])

driver = webdriver.Chrome(options=options)



wb = load_workbook('Products.xlsx')
ws = wb['Entries']
ws2 = wb['srno']


def Restart():
    driver.close()
    os.execv(sys.executable, ['python'] + sys.argv)





if __name__ == "__main__":

    row = 1
    col = 1
    ibtn = 1

    driver.get('https://supplier.meesho.com/')
    driver.maximize_window()

    start = time.time()
    Actions.Login(driver)
    end = time.time()
    t = end-start
    print("Time Taken : "+str(t))
    images = FetchingImages.ProductImages(ProImgDir)





    for x in images:
        try:

            OverAllTimeStarts = time.time()
            print('-----------------------------------------------------------------------------------')
            SlideImages = FetchingImages.SlideImages(SlideImagesPath)

            a = ws2['A1']
            n = int(a.value)


            AddSingleCatlog = driver.find_element(By.XPATH, '//*[@id="mainWrapper"]/div/div[1]/div[2]/div[1]/div[1]/div/button[2]')
            AddSingleCatlog.click()

            try:

                driver.implicitly_wait(7)
                Categories.HairDryer(driver)
            except Exception as e:
               driver.refresh()


            # Upload Product Front Image
            UploadProdImage = driver.find_element(By.XPATH, '//*[@id="addImages"]').send_keys(x)


            # time.sleep(5)
            driver.implicitly_wait(30)

            # Clicking to Continue button
            Continue = '//button[@class="MuiButton-root MuiButton-contained MuiButton-containedPrimary MuiButton-sizeSmall MuiButton-containedSizeSmall MuiButtonBase-root css-1n1ggul"]'
            ContinueBtn = driver.find_element(By.XPATH, Continue)
            ContinueBtn.click()

            try:
                wait = WebDriverWait(driver, 25)
                wait.until(EC.invisibility_of_element_located((By.XPATH, Continue)))
            except Exception as e:
                print("Exception : "+str(e))
                if ContinueBtn.text == 'Retry':
                    ContinueBtn.click()
                    wait = WebDriverWait(driver, 25)
                    wait.until(EC.invisibility_of_element_located((By.XPATH, Continue)))

            # time.sleep(8)
            driver.implicitly_wait(15)

            # i Button
            if ibtn <= 4:
                wait = WebDriverWait(driver, 30, ignored_exceptions=[NoSuchElementException])
                wait.until(EC.element_to_be_clickable((By.XPATH, '//p[normalize-space()="Wrong/Defective Returns Price"]//parent::div//parent::div/div[2]/div/div/div/div')))
                driver.find_element(By.XPATH, '//p[normalize-space()="Wrong/Defective Returns Price"]//parent::div//parent::div/div[2]/div/div/div/div').click()

            time.sleep(2)

            print(str(x)+' '+str(n))


            driver.execute_script("window.scrollBy(0,200)", "")

            driver.implicitly_wait(20)
            UploadSlideImages = driver.find_element(By.XPATH, '//input[@id="addMoreImagesInput"]')

            # Upload Product Slide Image
            for s in SlideImages:
                UploadSlideImages.send_keys(s)
                time.sleep(2)

            char1 = random.choice(range(65, 90))
            char2 = random.choice(range(65, 90))

            alpha = chr(char1)+"-"+chr(char2)


            start = time.time()
            Action.HairDryerInventoryDetails(driver, alpha)
            end = time.time()
            t = end-start
            print("TIme Taken : "+str(t))

            # driver.execute_script("arguments[0].scrollIntoView();", Fields.PackageDetail)


            start = time.time()
            Action.HairDryerProductDetails(driver)
            end = time.time()
            t = end-start
            print("Time Taken : "+str(t))



            # driver.execute_script("arguments[0].scrollIntoView();", Fields.Description)

            start = time.time()
            Action.HairDryerOtherAttributes(driver)
            end = time.time()
            t = end-start
            print("Time Taken : "+str(t))

            ws.cell(row=row, column=col).value = str(x) + ' ' + str(n)
            row += 1
            ibtn += 1


            SubmitButton = '//button[@class="MuiButton-root MuiButton-contained MuiButton-containedPrimary MuiButton-sizeMedium MuiButton-containedSizeMedium MuiButtonBase-root css-16z39kb"]'
            driver.find_element(By.XPATH, SubmitButton).click()
            time.sleep(1)
            driver.find_element(By.XPATH, '//button[@class="MuiButton-root MuiButton-contained MuiButton-containedPrimary MuiButton-sizeSmall MuiButton-containedSizeSmall MuiButtonBase-root css-1bc15yu"]').click()
            # os.replace(x, destination)
            # time.sleep(10)
            driver.implicitly_wait(15)
            print('Done')
            n += 1
            ws2['A1'] = n
            destination = path+'\\'+os.path.basename(x)
            os.replace(x, destination)
            print('File Name : ('+os.path.basename(x)+')\n has been moved to destination : '+destination)
            wb.save('Products.xlsx')
            OverAllTimeEnds = time.time()
            OverAllTime = OverAllTimeEnds-OverAllTimeStarts
            print('-----------------------------------------------------------------------------------')
            print("OverAll Time Taken : "+str(OverAllTime))
            print('-----------------------------------------------------------------------------------')
            print("\n")

        except (ElementClickInterceptedException, NoSuchElementException) as e:
            print("Exception : "+e)
            discardbtn = '//button[@class="MuiButton-root MuiButton-outlined MuiButton-outlinedPrimary MuiButton-sizeMedium MuiButton-outlinedSizeMedium MuiButtonBase-root css-1e66sag"]'
            process = '//button[@class="MuiButton-root MuiButton-contained MuiButton-containedPrimary MuiButton-sizeSmall MuiButton-containedSizeSmall MuiButtonBase-root css-1bc15yu"]'
            driver.find_element(By.XPATH, discardbtn).click()
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, process).click()
            driver.implicitly_wait(25)
            continue

ws2['A1'] = 1
print('Out of images')
wb.save('Products.xlsx')





